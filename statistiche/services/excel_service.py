import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


def genera_excel(titolo: str, colonne: list, righe: list) -> bytes:
    """Genera un file Excel formattato con i dati forniti."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titolo[:31]  # Excel limita a 31 caratteri

    n_col = len(colonne) if colonne else 1

    # Riga titolo (riga 1)
    ws.merge_cells(f'A1:{get_column_letter(n_col)}1')
    ws['A1'] = titolo
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1A1A2E')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Header colonne (riga 2)
    header_fill = PatternFill('solid', fgColor='D4A017')
    border = Border(
        bottom=Side(style='thin', color='999999'),
        right=Side(style='thin', color='CCCCCC'),
    )
    for col_idx, col_name in enumerate(colonne, 1):
        cell = ws.cell(row=2, column=col_idx, value=str(col_name))
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Dati con righe alternate (a partire dalla riga 3)
    fill_even = PatternFill('solid', fgColor='F5F0E8')
    for row_idx, row_data in enumerate(righe, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = fill_even

    # Autofit colonne
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                cell_len = len(str(cell.value or ''))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
